"""PDF to Excel conversion using PyMuPDF + openpyxl.

`find_tables()` can be slow on dense / scanned-looking pages so we run it
in a thread pool with a per-conversion timeout. openpyxl writes must
stay on a single thread (the Workbook isn't thread-safe), so parallelism
is limited to the table-extraction phase.
"""
from __future__ import annotations

import asyncio
import concurrent.futures
import logging
import os
import time
import uuid
from typing import Any

import fitz
from openpyxl import Workbook

from ..utils.cleanup import ensure_temp_dir, get_temp_path
from ..utils.exceptions import ToolTimeoutError

logger = logging.getLogger(__name__)

_CONVERSION_TIMEOUT_S = int(os.environ.get("PDF_TO_EXCEL_TIMEOUT_S", "60"))


def _extract_page(page) -> dict[str, Any]:
    """Pull tables + raw text off a single page in one shot.

    Returns a dict with `tables` (list of list-of-rows) and `lines`
    (list of stripped, non-empty strings). PyMuPDF releases the GIL on
    `find_tables()` / `get_text()` so this parallelises cleanly.
    """
    out: dict[str, Any] = {"tables": [], "lines": []}
    try:
        tables = page.find_tables()
        if tables and len(tables.tables) > 0:
            out["tables"] = [t.extract() for t in tables.tables]
            return out
    except Exception:
        pass

    text = page.get_text("text") or ""
    out["lines"] = [line.strip() for line in text.split("\n") if line.strip()]
    return out


def _build_workbook(input_path: str) -> tuple[str, bool]:
    """Synchronous workbook build. Returns (output_path, found_anything)."""
    doc = fitz.open(input_path)
    try:
        page_count = len(doc)

        # Parallelise the heavy extraction work — keep the workbook write
        # sequential after to honour openpyxl's single-thread requirement.
        payloads: list[dict[str, Any]] = [None] * page_count  # type: ignore[list-item]
        if page_count > 1:
            max_workers = max(1, (os.cpu_count() or 2) - 1)
            max_workers = min(max_workers, 4, page_count)
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {pool.submit(_extract_page, doc[i]): i for i in range(page_count)}
                for fut in concurrent.futures.as_completed(futures):
                    payloads[futures[fut]] = fut.result()
        elif page_count == 1:
            payloads[0] = _extract_page(doc[0])

        wb = Workbook()
        ws = wb.active
        ws.title = "Page 1"

        found_anything = False
        for page_num in range(page_count):
            if page_num > 0:
                ws = wb.create_sheet(title=f"Page {page_num + 1}")
            row = 1
            payload = payloads[page_num] or {}
            tables = payload.get("tables") or []
            if tables:
                for table in tables:
                    for trow in table:
                        for col, cell in enumerate(trow, 1):
                            ws.cell(row=row, column=col, value=cell or "")
                        row += 1
                    row += 1
                found_anything = True
                continue

            for line in payload.get("lines", []):
                ws.cell(row=row, column=1, value=line)
                row += 1
                found_anything = True
    finally:
        doc.close()

    if not found_anything:
        # Image-only PDF — no text and no detectable tables. Surface the
        # friendly error so users know to OCR first instead of staring at
        # an empty workbook.
        raise ValueError(
            "No tables detected. PDF may be a scan — try OCR first to make it searchable"
        )
    output_path = get_temp_path(f"converted_{uuid.uuid4().hex}.xlsx")
    wb.save(str(output_path))
    return (str(output_path), found_anything)


async def convert_to_excel(input_path: str) -> str:
    ensure_temp_dir()
    started = time.monotonic()
    try:
        input_size = os.path.getsize(input_path)
    except OSError:
        input_size = 0
    logger.info("pdf_to_excel: start input_bytes=%d", input_size)

    # Run the synchronous build in a thread so the event loop can cancel
    # it via asyncio.wait_for. PyMuPDF + openpyxl both release the GIL
    # for their heavy lifting so this doesn't starve the loop.
    loop = asyncio.get_running_loop()
    try:
        output_path, _ = await asyncio.wait_for(
            loop.run_in_executor(None, _build_workbook, input_path),
            timeout=_CONVERSION_TIMEOUT_S,
        )
    except asyncio.TimeoutError as exc:
        logger.warning(
            "pdf_to_excel: timeout after %ds input_bytes=%d",
            _CONVERSION_TIMEOUT_S, input_size,
        )
        raise ToolTimeoutError(
            f"Conversion took longer than {_CONVERSION_TIMEOUT_S}s — try a smaller PDF"
        ) from exc

    duration_ms = int((time.monotonic() - started) * 1000)
    try:
        out_size = os.path.getsize(output_path)
    except OSError:
        out_size = 0
    logger.info(
        "pdf_to_excel: done duration_ms=%d input_bytes=%d output_bytes=%d",
        duration_ms, input_size, out_size,
    )
    return output_path
