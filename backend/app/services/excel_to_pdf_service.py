import uuid
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from ..utils.cleanup import get_temp_path, ensure_temp_dir


def excel_to_pdf(input_path: str) -> str:
    """Convert .xlsx file to PDF using openpyxl + reportlab."""
    ensure_temp_dir()
    output_path = get_temp_path(f"excel_{uuid.uuid4().hex}.pdf")

    wb = load_workbook(input_path, data_only=True)
    c = canvas.Canvas(str(output_path), pagesize=landscape(A4))
    page_width, page_height = landscape(A4)
    margin = 36  # 0.5 inch

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Calculate column widths based on content
        num_cols = max(len(row) for row in rows)
        if num_cols == 0:
            continue

        usable_width = page_width - 2 * margin
        col_width = usable_width / num_cols
        row_height = 16
        font_size = 8
        y = page_height - margin

        # Sheet title
        c.setFont("Helvetica-Bold", 12)
        c.drawString(margin, y, sheet.title)
        y -= 20

        c.setFont("Helvetica", font_size)

        for row_idx, row in enumerate(rows):
            if y < margin + row_height:
                c.showPage()
                c.setFont("Helvetica", font_size)
                y = page_height - margin

            # Draw header row with bold
            if row_idx == 0:
                c.setFont("Helvetica-Bold", font_size)
                # Header background
                c.setFillColorRGB(0.9, 0.9, 0.95)
                c.rect(margin, y - 4, usable_width, row_height, fill=1, stroke=0)
                c.setFillColorRGB(0, 0, 0)

            x = margin
            for col_idx, cell in enumerate(row):
                if col_idx >= num_cols:
                    break
                text = str(cell) if cell is not None else ""
                # Truncate if too long
                max_chars = int(col_width / (font_size * 0.5))
                if len(text) > max_chars:
                    text = text[:max_chars - 1] + "…"
                c.drawString(x + 3, y, text)
                x += col_width

            # Draw grid lines
            c.setStrokeColorRGB(0.8, 0.8, 0.8)
            c.line(margin, y - 4, margin + usable_width, y - 4)

            if row_idx == 0:
                c.setFont("Helvetica", font_size)

            y -= row_height

        c.showPage()

    c.save()
    wb.close()
    return str(output_path)
